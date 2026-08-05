"""Excel export.

**Data first, charts second.** A spreadsheet is the one format a recipient
will actually work in — re-sort it, pivot it, paste it into their own model —
so every table lands as a real worksheet range with typed cells, and charts
are native Excel objects anchored to those ranges rather than pictures. An
analyst can change a number and watch the chart move, which is the entire
reason they asked for Excel rather than the PDF.

Numbers are written as numbers. Writing "1,234,567" as a string produces a
workbook that looks right and cannot be summed, and the person who discovers
that is the one who needed the total.
"""

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.reporting.contracts import (
    Block,
    BlockKind,
    ChartKind,
    ExportFormat,
    Renderer,
    Report,
    Section,
)

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F3864")
NOTE_FONT = Font(size=9, italic=True, color="595959")

CURRENCY_FORMAT = "#,##0"
RATE_FORMAT = "0.0%"

#: Excel forbids these in sheet names, and silently truncates past 31 chars.
_ILLEGAL_SHEET_CHARS = set(r"[]:*?/\\")


class XlsxRenderer(Renderer):
    format = ExportFormat.XLSX
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    extension = "xlsx"

    def render(self, report: Report) -> bytes:
        workbook = Workbook()
        workbook.remove(workbook.active)  # type: ignore[arg-type]

        self._cover(workbook, report)
        for section in report.sections:
            self._section(workbook, section)

        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    # ── Sheets ───────────────────────────────────────────────────────

    def _cover(self, workbook: Workbook, report: Report) -> None:
        sheet = workbook.create_sheet("Overview")
        sheet["A1"] = report.title
        sheet["A1"].font = Font(size=18, bold=True, color="1F3864")
        sheet["A2"] = report.subtitle
        sheet["A3"] = report.period_label
        sheet["A4"] = f"Generated {report.generated_at:%Y-%m-%d %H:%M UTC}"

        row = 6
        sheet.cell(row=row, column=1, value="Caveats").font = TITLE_FONT
        for caveat in report.caveats:
            row += 1
            cell = sheet.cell(row=row, column=1, value=caveat)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions["A"].width = 110

    def _section(self, workbook: Workbook, section: Section) -> None:
        sheet = workbook.create_sheet(_sheet_name(section.title))
        row = 1

        sheet.cell(row=row, column=1, value=section.title).font = TITLE_FONT
        row += 1
        if section.subtitle:
            sheet.cell(row=row, column=1, value=section.subtitle).font = NOTE_FONT
            row += 1
        if section.unavailable_reason:
            # An empty section carries its reason rather than being dropped:
            # "nothing found" and "never ran" are opposite conclusions.
            sheet.cell(row=row, column=1, value=section.unavailable_reason).font = NOTE_FONT
            sheet.column_dimensions["A"].width = 110
            return

        row += 1
        for block in section.blocks:
            row = self._block(sheet, block, row) + 2

        sheet.column_dimensions["A"].width = 46
        for index in range(2, 8):
            sheet.column_dimensions[get_column_letter(index)].width = 18

    def _block(self, sheet: Worksheet, block: Block, row: int) -> int:
        if block.title:
            sheet.cell(row=row, column=1, value=block.title).font = Font(bold=True)
            row += 1

        if block.kind is BlockKind.NARRATIVE and block.text:
            cell = sheet.cell(row=row, column=1, value=block.text)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            return row

        if block.kind is BlockKind.CALLOUT:
            for bullet in block.bullets:
                cell = sheet.cell(row=row, column=1, value=f"• {bullet}")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
            return row - 1

        if block.kind is BlockKind.KPI_GRID:
            return self._kpis(sheet, block, row)

        if block.rows:
            return self._table(sheet, block, row)

        return row

    def _kpis(self, sheet: Worksheet, block: Block, row: int) -> int:
        headers = ("Measure", "Value", "Prior", "Change")
        for column, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row, column=column, value=header)
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        row += 1

        for kpi in block.kpis:
            sheet.cell(row=row, column=1, value=kpi.label)
            value_cell = sheet.cell(row=row, column=2, value=kpi.value)
            value_cell.number_format = RATE_FORMAT if kpi.unit == "rate" else CURRENCY_FORMAT
            if kpi.comparison is not None:
                prior = sheet.cell(row=row, column=3, value=kpi.comparison)
                prior.number_format = value_cell.number_format
            if kpi.change is not None:
                change = sheet.cell(row=row, column=4, value=kpi.change)
                change.number_format = RATE_FORMAT
            row += 1

        if block.note:
            sheet.cell(row=row, column=1, value=block.note).font = NOTE_FONT
        return row

    def _table(self, sheet: Worksheet, block: Block, row: int) -> int:
        header_row = row
        for column, name in enumerate(block.columns, start=1):
            cell = sheet.cell(row=row, column=column, value=name.replace("_", " ").title())
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        row += 1

        first_data_row = row
        for record in block.rows:
            for column, value in enumerate(record, start=1):
                cell = sheet.cell(row=row, column=column, value=value)
                # Typed, not stringified. A workbook full of text that looks
                # numeric is a workbook nobody can total.
                if isinstance(value, int | float):
                    cell.number_format = CURRENCY_FORMAT
            row += 1

        if block.chart_kind and block.rows:
            self._chart(sheet, block, header_row, first_data_row, row - 1)
            row += 16  # leave room for the chart anchored below

        if block.note:
            sheet.cell(row=row, column=1, value=block.note).font = NOTE_FONT
        return row

    def _chart(
        self, sheet: Worksheet, block: Block, header_row: int, first: int, last: int
    ) -> None:
        """Anchor a native chart to the range just written.

        Native rather than an image: the recipient can extend the range, change
        a value, or restyle it. A picture of a chart is a dead end.
        """
        try:
            category_index = block.columns.index(block.chart_category_column) + 1
        except ValueError:
            category_index = 1

        chart: Any = LineChart() if block.chart_kind is ChartKind.LINE else BarChart()
        if block.chart_kind is ChartKind.HORIZONTAL_BAR:
            chart.type = "bar"
        elif block.chart_kind is ChartKind.BAR:
            chart.type = "col"
        chart.title = block.title or "Chart"
        chart.height, chart.width = 8, 18

        categories = Reference(sheet, min_col=category_index, min_row=first, max_row=last)
        for name in block.chart_value_columns:
            if name not in block.columns:
                continue
            index = block.columns.index(name) + 1
            data = Reference(sheet, min_col=index, min_row=header_row, max_row=last)
            chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        sheet.add_chart(chart, f"{get_column_letter(len(block.columns) + 2)}{first}")


def _sheet_name(title: str) -> str:
    """Excel sheet names: 31 characters, no []:*?/\\ ."""
    cleaned = "".join(char for char in title if char not in _ILLEGAL_SHEET_CHARS)
    return cleaned[:31] or "Sheet"
